#!/usr/bin/env python3
"""
Script за импорт на Регулативи од Excel во JSON формат
Извор: kb/Raw_Files/Spisok na Regulativi KN 15.xlsx
Output: kb/processed/regulations_data.json
"""

import openpyxl
import json
import sys
import re
from pathlib import Path
from datetime import datetime

def parse_date(date_str):
    """Парсирај датум од формат 'број/година'"""
    if not date_str:
        return None
    try:
        # Формат: "222/2020\n17.09.2020"
        if '\n' in date_str:
            date_part = date_str.split('\n')[1].strip()
            return datetime.strptime(date_part, '%d.%m.%Y').isoformat()
        return None
    except:
        return None

def clean_tariff(tariff_str):
    """Почисти тарифна ознака (отстрани spaces)"""
    if not tariff_str:
        return None
    return tariff_str.replace(' ', '').strip()

def import_regulations():
    """Импортира регулативи од Excel"""
    
    # Патеки
    excel_file = Path(__file__).parent.parent / "Raw_Files" / "Spisok na Regulativi KN 15.xlsx"
    output_file = Path(__file__).parent.parent / "processed" / "regulations_data.json"
    
    # Креирај processed фолдер
    output_file.parent.mkdir(exist_ok=True)
    
    print(f"📂 Отварање: {excel_file}")
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb['Sheet1']
    
    print(f"📊 Вкупно редови: {ws.max_row}")
    
    # Прочитај хедери
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    
    # Мапирање на колони
    regulation_records = []
    skipped = 0
    
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            # Извлечи податоци
            celex = str(row[0].value or "").strip()
            
            # Прескокни ако нема CELEX
            if not celex:
                skipped += 1
                continue
            
            gazette_ref = str(row[1].value or "").strip()
            desc_en = str(row[2].value or "").strip()
            desc_mk = str(row[3].value or "").strip()
            legal_basis = str(row[4].value or "").strip()
            tariff = clean_tariff(str(row[5].value or "").strip())
            
            record = {
                "celexNumber": celex,
                "officialGazetteRef": gazette_ref if gazette_ref else None,
                "tariffNumber": tariff,
                "descriptionEN": desc_en if desc_en else None,
                "descriptionMK": desc_mk if desc_mk else None,
                "legalBasis": legal_basis if legal_basis else None,
                "effectiveDate": parse_date(gazette_ref),
                "isActive": True
            }
            
            regulation_records.append(record)
            
            # Progress
            if i % 200 == 0:
                print(f"  ⏳ Процесирани: {i-1} редови...")
        
        except Exception as e:
            print(f"  ⚠️  Грешка на ред {i}: {e}")
            skipped += 1
            continue
    
    wb.close()
    
    # Зачувај во JSON
    print(f"\n💾 Зачувување во: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(regulation_records, f, ensure_ascii=False, indent=2)
    
    # Статистика
    print(f"\n✅ Успешно импортирани: {len(regulation_records)} записи")
    print(f"⚠️  Прескокнати: {skipped} записи")
    print(f"📦 Фајл: {output_file}")
    print(f"📏 Големина: {output_file.stat().st_size / 1024:.2f} KB")
    
    # Примери
    print(f"\n📋 Примери (прва 3 записи):")
    for rec in regulation_records[:3]:
        print(f"  {rec['celexNumber']}: Тарифа {rec['tariffNumber']}")
        print(f"    МК: {rec['descriptionMK'][:60]}...")

if __name__ == "__main__":
    try:
        import_regulations()
    except FileNotFoundError as e:
        print(f"❌ Грешка: Фајлот не постои: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Грешка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
