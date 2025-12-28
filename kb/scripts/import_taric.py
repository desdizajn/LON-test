#!/usr/bin/env python3
"""
Script за импорт на TARIC податоци од Excel во JSON формат
Извор: kb/Raw_Files/TARIC.xlsx
Output: kb/processed/taric_data.json
"""

import openpyxl
import json
import sys
from pathlib import Path

def import_taric():
    """Импортира TARIC податоци од Excel"""
    
    # Патеки
    excel_file = Path(__file__).parent.parent / "Raw_Files" / "TARIC.xlsx"
    output_file = Path(__file__).parent.parent / "processed" / "taric_data.json"
    
    # Креирај processed фолдер
    output_file.parent.mkdir(exist_ok=True)
    
    print(f"📂 Отварање: {excel_file}")
    wb = openpyxl.load_workbook(excel_file, read_only=True, data_only=True)
    ws = wb['Real life data']
    
    print(f"📊 Вкупно редови: {ws.max_row}")
    
    # Прочитај хедери
    headers = [cell.value for cell in next(ws.iter_rows(max_row=1))]
    print(f"📋 Колони: {', '.join(headers[:8])}")
    
    # Мапирање на колони
    tariff_records = []
    skipped = 0
    
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        try:
            # Извлечи податоци
            tariff_number = str(row[0].value).strip() if row[0].value else None
            
            # Прескокни ако нема тарифна ознака
            if not tariff_number or len(tariff_number) != 10:
                skipped += 1
                continue
            
            record = {
                "tariffNumber": tariff_number,
                "tarbr": str(row[1].value or "").strip(),
                "taroz1": str(row[2].value or "").strip(),
                "taroz2": str(row[3].value or "").strip(),
                "taroz3": str(row[4].value or "").strip(),
                "description": str(row[5].value or "").strip(),
                "customsRate": float(row[6].value) if row[6].value else None,
                "unitMeasure": str(row[7].value or "").strip() if row[7].value else None,
                "fi": str(row[8].value or "").strip() if row[8].value else None,
                "fu": str(row[9].value or "").strip() if row[9].value else None,
                "pv": str(row[10].value or "").strip() if row[10].value else None,
                "vatRate": float(row[18].value) if row[18].value else None,
                "ex": str(row[19].value or "").strip() if row[19].value else None,
                "isActive": True
            }
            
            tariff_records.append(record)
            
            # Progress
            if i % 1000 == 0:
                print(f"  ⏳ Процесирани: {i-1} редови...")
        
        except Exception as e:
            print(f"  ⚠️  Грешка на ред {i}: {e}")
            skipped += 1
            continue
    
    wb.close()
    
    # Зачувај во JSON
    print(f"\n💾 Зачувување во: {output_file}")
    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(tariff_records, f, ensure_ascii=False, indent=2)
    
    # Статистика
    print(f"\n✅ Успешно импортирани: {len(tariff_records)} записи")
    print(f"⚠️  Прескокнати: {skipped} записи")
    print(f"📦 Фајл: {output_file}")
    print(f"📏 Големина: {output_file.stat().st_size / 1024 / 1024:.2f} MB")
    
    # Примери
    print(f"\n📋 Примери (прва 3 записи):")
    for rec in tariff_records[:3]:
        print(f"  {rec['tariffNumber']}: {rec['description'][:60]}")

if __name__ == "__main__":
    try:
        import_taric()
    except FileNotFoundError as e:
        print(f"❌ Грешка: Фајлот не постои: {e}")
        sys.exit(1)
    except Exception as e:
        print(f"❌ Грешка: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
